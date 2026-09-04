"""Прямая загрузка страниц и документов — без браузера.

Зачем это нужно. Раньше единственным способом посмотреть, что лежит по
адресу, был `browser_exec`: открыть вкладку, поспать несколько секунд,
вытащить `document.body.innerText`. Один адрес — один шаг агента. В разборе
логов видно, чем это кончается: десять шагов подряд уходит на проверку
адресов, которых просто не существует (NXDOMAIN, 404), и до полезной работы
дело не доходит.

Здесь адреса берутся пачкой и параллельно, ответ приходит за доли секунды, а
не за десять секунд сна, и PDF/Excel/JSON разбираются сразу, без отдельного
шага на скачивание. Браузер остаётся для того, ради чего он и нужен: страниц,
которые собираются скриптами, и кликов.
"""

import asyncio
import csv
import io
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

import httpx
from bs4 import BeautifulSoup

from app.config import config
from app.logger import logger
from pydantic import Field

from app.tool.base import BaseTool, ToolResult


MAX_URLS = 8
TIMEOUT = 25.0
DEFAULT_MAX_CHARS = 20_000
# больше этого в память не берём: файл уедет на диск, а агент прочитает его сам
MAX_BYTES = 40 * 1024 * 1024

# Заголовки обычного браузера. Многие сайты отдают 403 на клиент без них —
# это не обход защиты, а всего лишь отказ от представления «я скрипт».
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9,ru;q=0.8,hy;q=0.7",
}

# Что делать с источником, который прямым запросом не открылся. Формулировка
# намеренно предписывающая и с конкретным адресом: в разобранном прогоне общий
# совет «попробуйте browser_exec» был проигнорирован четыре раза подряд, и
# закрытый источник молча заменялся пересказом из поисковой выдачи.
_ESCALATE = (
    "\nСЛЕДУЮЩЕЕ ДЕЙСТВИЕ: откройте этот же адрес через browser_exec:\n"
    "    new_tab(\"{url}\"); wait_for_load(); print(js(\"document.body.innerText\"))\n"
    "Не подменяйте этот источник пересказом из поисковой выдачи и не объявляйте "
    "его недоступным, пока браузер не попробовал. Если и браузер не справится — "
    "так и запишите, назвав оба способа."
)

# Щит от автоматики — отдельный случай. Ломиться в него бессмысленно и
# незачем: это чужое решение о том, кого пускать. Одна попытка браузером
# оправдана (иногда щит пропускает настоящий браузер), но готовиться надо к
# отказу и сразу искать данные там, где их отдают добровольно.
_SHIELD_PLAN = (
    "\nЧТО ДЕЛАТЬ: одна попытка через browser_exec допустима —\n"
    "    new_tab(\"{url}\"); wait_for_load(); print(js(\"document.body.innerText\"))\n"
    "Но рассчитывайте на отказ и параллельно ищите замену. Щит не обходят "
    "повторами, сменой заголовков или регистрацией — это решение владельца "
    "сайта о том, кого пускать. Ищите те же сведения там, где их публикуют "
    "сами: статистические органы, кадастр, регулятор, открытые данные, "
    "выгрузки в XLSX/CSV. Данные о состоявшихся сделках обычно и точнее, "
    "чем объявления с запрашиваемыми ценами. Если замены нет — запишите в "
    "находки, что источник закрыт щитом, и что именно из-за этого неизвестно."
)

# Дверь с замком — третий случай, отличный и от щита, и от каприза заголовков.
# Здесь у человека за интерфейсом доступ, скорее всего, есть, и он может им
# поделиться, не отдавая пароль.
_LOGIN_PLAN = (
    "\nПОХОЖЕ НА ТРЕБОВАНИЕ ВОЙТИ. Если эти данные важны для задачи и у "
    "человека, вероятно, есть сюда доступ (банк, платная подписка, кабинет "
    "регулятора) — вызовите request_login с этим адресом и объясните, какие "
    "сведения нужны. Работа встанет на паузу, человек войдёт своими руками, "
    "после чего откройте страницу заново через browser_exec. Если источник "
    "второстепенный — не отвлекайте человека, запишите пробел и идите дальше."
)

# теги, чей текст не имеет отношения к содержанию страницы
NOISE = ("script", "style", "noscript", "template", "svg", "iframe")


def _clean(text: str) -> str:
    """Схлопывает пустые строки: страницы дают их сотнями."""
    text = re.sub(r"[ \t\xa0]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _from_html(raw: bytes, encoding: Optional[str]) -> str:
    soup = BeautifulSoup(raw.decode(encoding or "utf-8", "replace"), "html.parser")
    for tag in soup(NOISE):
        tag.decompose()
    title = soup.title.get_text(strip=True) if soup.title else ""
    body = soup.get_text("\n")
    return _clean((f"# {title}\n\n" if title else "") + body)


def _from_pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:  # pragma: no cover - зависит от сборки образа
        return "[PDF: не установлен pypdf, текст не извлечён]"
    reader = PdfReader(io.BytesIO(raw))
    pages = []
    for number, page in enumerate(reader.pages, 1):
        try:
            pages.append(f"\n--- страница {number} ---\n{page.extract_text() or ''}")
        except Exception as error:  # одна битая страница не должна ронять весь файл
            pages.append(f"\n--- страница {number}: не читается ({error}) ---")
    return _clean("".join(pages))


def _from_xlsx(raw: bytes) -> str:
    """Карта книги: листы, размеры, заголовки и первые строки.

    Полная выгрузка здесь не нужна и вредна — она забьёт контекст. Задача
    инструмента: показать агенту структуру, чтобы он точно знал, какой лист и
    какие колонки читать дальше через python_execute.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        return "[XLSX: не установлен openpyxl]"
    book = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    parts = [f"Листы: {book.sheetnames}"]
    for name in book.sheetnames:
        sheet = book[name]
        parts.append(f"\n=== лист «{name}» — строк {sheet.max_row}, колонок {sheet.max_column} ===")
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 8), values_only=True):
            cells = ["" if v is None else str(v)[:28] for v in row[:16]]
            parts.append(" | ".join(cells))
    return _clean("\n".join(parts))


def _from_csv(raw: bytes, encoding: Optional[str]) -> str:
    text = raw.decode(encoding or "utf-8", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4000])
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    head = "\n".join(" | ".join(cell[:28] for cell in row[:16]) for row in rows[:25])
    return _clean(f"Строк: {len(rows)}, разделитель: {delimiter!r}\n{head}")


def _from_json(raw: bytes, encoding: Optional[str]) -> str:
    try:
        data = json.loads(raw.decode(encoding or "utf-8", "replace"))
    except ValueError as error:
        return f"[не разобрался как JSON: {error}]"
    return json.dumps(data, ensure_ascii=False, indent=2)


def _extension(url: str, content_type: str) -> str:
    for marker, suffix in (
        ("pdf", ".pdf"),
        ("spreadsheet", ".xlsx"),
        ("ms-excel", ".xls"),
        ("zip", ".zip"),
        ("msword", ".doc"),
        ("wordprocessing", ".docx"),
    ):
        if marker in content_type:
            return suffix
    return Path(urlparse(url).path).suffix or ".bin"


def _safe_name(url: str, content_type: str) -> str:
    stem = Path(urlparse(url).path).stem or urlparse(url).netloc.replace(".", "_")
    stem = re.sub(r"[^A-Za-z0-9_.-]", "_", stem)[:60] or "download"
    suffix = _extension(url, content_type)
    return stem + (suffix if not stem.endswith(suffix) else "")


class Fetch(BaseTool):
    """Загрузка адресов напрямую, без браузера."""

    name: str = "fetch"
    description: str = (
        "Fetch one or several URLs directly over HTTP and return their readable "
        "content. Much faster and cheaper than the browser: pass up to 8 URLs at "
        "once and they are fetched in parallel, with no page load waits.\n"
        "Understands HTML (text extraction), PDF, XLSX/XLS, CSV, JSON, XML and "
        "plain text. Binary documents are also saved into the working directory "
        "so python_execute can analyse them further; the saved path is reported.\n"
        "USE THIS FIRST for: checking whether a URL exists, reading articles, "
        "documentation, statistics releases, API endpoints, and downloading "
        "reports. Only fall back to browser_exec when the page needs JavaScript "
        "to render, needs clicking/typing, or when fetch reports it was blocked."
    )
    # Папка задачи. Интерфейс подставляет сюда workspace/task_<id>, чтобы
    # скачанные отчёты лежали рядом с остальными файлами этой задачи и
    # удалялись вместе с ней.
    directory: str = ""

    # Адреса, которые прямым запросом не даются: 403, антибот, страница
    # собирается скриптами. Их берёт браузер. В разборе прогона агент такие
    # источники просто бросал и подменял поисковой выдачей — по list.am,
    # главной доске объявлений Армении, так потерялись все прямые данные.
    # Список читает агент перед завершением шага, см. app/agent/manus.py.
    blocked_urls: List[str] = Field(default_factory=list)

    parameters: dict = {
        "type": "object",
        "properties": {
            "urls": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    f"(required) 1 to {MAX_URLS} URLs to fetch in parallel. "
                    "Probing several candidate URLs in one call is the intended "
                    "usage — it costs the same as probing one."
                ),
            },
            "max_chars": {
                "type": "integer",
                "description": (
                    f"(optional) Characters of text to return per URL. "
                    f"Default {DEFAULT_MAX_CHARS}."
                ),
            },
            "probe_only": {
                "type": "boolean",
                "description": (
                    "(optional) Only report status code, content type and size "
                    "without returning content. Use to check quickly which of "
                    "many candidate URLs actually exist."
                ),
            },
        },
        "required": ["urls"],
    }

    async def execute(
        self,
        urls: List[str],
        max_chars: int = DEFAULT_MAX_CHARS,
        probe_only: bool = False,
        **kwargs: Any,
    ) -> ToolResult:
        if isinstance(urls, str):
            urls = [urls]
        urls = [u.strip() for u in urls if isinstance(u, str) and u.strip()][:MAX_URLS]
        if not urls:
            return ToolResult(error="Не передано ни одного адреса.")

        max_chars = max(500, min(int(max_chars or DEFAULT_MAX_CHARS), 100_000))

        async with httpx.AsyncClient(
            follow_redirects=True,
            timeout=TIMEOUT,
            headers=HEADERS,
            verify=True,
        ) as client:
            results = await asyncio.gather(
                *(self._one(client, url, max_chars, probe_only) for url in urls)
            )

        return ToolResult(output="\n\n".join(results))

    async def _one(
        self,
        client: httpx.AsyncClient,
        url: str,
        max_chars: int,
        probe_only: bool,
    ) -> str:
        if not urlparse(url).scheme:
            url = "https://" + url
        head = f"===== {url}"
        try:
            response = await client.get(url)
        except httpx.HTTPError as error:
            # Отдельно называем несуществующий домен: модель по логам склонна
            # угадывать адреса, и ей полезно видеть разницу между «сайта нет»
            # и «сайт есть, но не пустил».
            reason = _explain(error)
            return f"{head}\nНЕ ОТКРЫЛСЯ: {reason}"

        content_type = (response.headers.get("content-type") or "").lower()
        raw = response.content
        landed = str(response.url)
        head = f"===== {landed}"
        if landed != url:
            head += f"\n(перенаправление с {url})"
        head += f"\nстатус {response.status_code} · {content_type or 'тип не указан'} · {len(raw)} байт"

        if response.status_code >= 400:
            hint = _blocked_hint(response.status_code, raw, content_type)
            if response.status_code in (401, 403, 429):
                self._remember_blocked(landed)
                if _shield_in(raw):
                    hint += _SHIELD_PLAN.format(url=landed)
                elif response.status_code == 401:
                    hint += _LOGIN_PLAN
                else:
                    hint += _ESCALATE.format(url=landed)
            return f"{head}\nСТРАНИЦА НЕ ОТДАНА.{hint}"
        if probe_only:
            return f"{head}\n(запрошена только проверка адреса)"
        if len(raw) > MAX_BYTES:
            return f"{head}\nСлишком большой файл, не читаю."

        saved = self._save_if_binary(landed, content_type, raw)
        try:
            body = _render(raw, content_type, landed, response.encoding)
        except Exception as error:  # разбор не должен ронять весь вызов
            logger.warning(f"fetch: не разобрал {landed}: {error}")
            body = f"[содержимое не разобрано: {error}]"

        # Признаки страницы считаем ДО обрезки: иначе меряем длину среза,
        # а не длину содержимого, и любая страница выглядит пустой.
        blocked = _looks_blocked(body)
        script_built = _looks_script_built(raw, body, content_type)
        # Стена входа обычно отвечает кодом 200: сервер честно отдал страницу,
        # просто это форма входа, а не данные. Ни щитом, ни пустым каркасом
        # такое не поймать — нужен отдельный разбор.
        login_wall = looks_like_login(landed, body, response.status_code)

        if len(body) > max_chars:
            body = body[:max_chars] + f"\n\n[…обрезано, всего {len(body)} символов]"
        if saved:
            head += f"\nсохранено в файл: {saved}"
        if blocked:
            self._remember_blocked(landed)
            head += (
                f"\nПОХОЖЕ НА ЗАЩИТУ ОТ БОТОВ ({blocked}) — прямым запросом "
                "этот источник не взять." + _ESCALATE.format(url=landed)
            )
        elif login_wall:
            self._remember_blocked(landed)
            head += _LOGIN_PLAN
        elif script_built:
            self._remember_blocked(landed)
            head += (
                "\nТЕКСТА ПОЧТИ НЕТ: страница весит много, а читаемого текста "
                "мало — содержимое рисуют скрипты, ниже только меню."
                + _ESCALATE.format(url=landed)
            )
        return f"{head}\n\n{body}"

    def _remember_blocked(self, url: str) -> None:
        if url not in self.blocked_urls:
            self.blocked_urls.append(url)

    def _save_if_binary(self, url: str, content_type: str, raw: bytes) -> Optional[str]:
        """Документы кладём в рабочую папку: они ещё понадобятся python_execute."""
        binary = any(
            marker in content_type
            for marker in ("pdf", "spreadsheet", "ms-excel", "zip", "msword", "wordprocessing")
        )
        if not binary:
            return None
        try:
            folder = Path(self.directory or config.workspace_root)
            folder.mkdir(parents=True, exist_ok=True)
            path = folder / _safe_name(url, content_type)
            path.write_bytes(raw)
            return str(path)
        except OSError as error:
            logger.warning(f"fetch: не сохранил {url}: {error}")
            return None


def _render(raw: bytes, content_type: str, url: str, encoding: Optional[str]) -> str:
    lowered = url.lower()
    if "pdf" in content_type or lowered.endswith(".pdf"):
        return _from_pdf(raw)
    if "spreadsheet" in content_type or lowered.endswith((".xlsx", ".xlsm")):
        return _from_xlsx(raw)
    if "json" in content_type or lowered.endswith(".json"):
        return _from_json(raw, encoding)
    if "csv" in content_type or lowered.endswith((".csv", ".tsv")):
        return _from_csv(raw, encoding)
    if "html" in content_type or "xhtml" in content_type:
        return _from_html(raw, encoding)
    if content_type.startswith("text/") or "xml" in content_type:
        return _clean(raw.decode(encoding or "utf-8", "replace"))
    if not content_type:
        # тип не указан — пробуем как текст, для самодельных API это обычное дело
        return _clean(raw.decode(encoding or "utf-8", "replace"))
    return f"[двоичные данные {content_type}, {len(raw)} байт — читайте файл через python_execute]"


def _explain(error: httpx.HTTPError) -> str:
    text = str(error)
    if isinstance(error, httpx.ConnectError) and (
        "Name or service not known" in text or "nodename nor servname" in text
    ):
        return "такого домена не существует (DNS не разрешается). Адрес выдуман — не угадывайте его снова, найдите через поиск."
    if isinstance(error, httpx.TimeoutException):
        return "сервер не ответил за отведённое время"
    return f"{type(error).__name__}: {text}"


# Как выглядит стена входа. Держим правило узким: лучше не заметить дверь,
# чем принять за дверь обычную страницу с формой подписки в подвале.
_FORM = ('type="password"', "type='password'", 'name="password"')
_WORDS = (
    "sign in", "log in", "login", "войти", "вход", "авторизац",
    "մուտք",  # армянский «вход» — целевой рынок этого пользователя
)
_PATHS = ("/login", "/signin", "/sign-in", "/auth", "/account/login", "/user/login")


def looks_like_login(final_url: str, body: str, status: int) -> bool:
    """Похоже ли, что нас развернули на форму входа.

    Требуем два признака сразу: поле пароля и слово о входе рядом. Одного
    поля мало — форма входа в подвале есть у половины сайтов, и принимать
    каждую такую страницу за стену значило бы дёргать человека впустую.
    """
    if status == 401:
        return True
    lowered = body[:20_000].lower()
    has_field = any(marker in lowered for marker in _FORM)
    has_word = any(word in lowered for word in _WORDS)
    path = urlparse(final_url).path.lower()
    redirected = any(path.startswith(p) or path == p.strip("/") for p in _PATHS)
    return (has_field and has_word and redirected) or (has_field and redirected)


def _blocked_hint(status: int, raw: bytes, content_type: str) -> str:
    if status in (401, 403):
        # Отказ отказу рознь, и реакция на них разная. Щит вроде Cloudflare
        # срабатывает ДО всякого входа: личный кабинет тут не поможет, и
        # браузер чаще всего тоже. Обычный же 403 бывает капризом заголовков,
        # и браузер его нередко снимает. Агенту нужно знать, в каком он случае.
        if _shield_in(raw):
            return (
                " Это щит от автоматики (Cloudflare или подобный), а НЕ требование "
                "войти в личный кабинет: проверка срабатывает раньше формы входа, "
                "поэтому учётная запись здесь ничего не даст."
            )
        return " Доступ закрыт (403/401): нужен вход или сайт не пускает автоматику."
    if status == 404:
        return " Такой страницы нет. Не подбирайте адрес вручную — найдите ссылку поиском или в карте сайта."
    if status == 429:
        return " Слишком много запросов, сайт просит подождать."
    if status >= 500:
        return " Ошибка на стороне сайта, не в запросе."
    return ""


BOT_WALL = (
    ("cloudflare", "Cloudflare"),
    ("checking your browser", "Cloudflare"),
    ("security verification", "Cloudflare"),
    ("unusual traffic", "антибот Google"),
    ("captcha", "капча"),
    ("are you a robot", "капча"),
    ("enable javascript and cookies", "антибот"),
)


# Пороги подобраны по замерам на живых страницах, а не на глаз:
#   страница ЦБА, где таблицу рисует скрипт — 3.0% текста, 1543 знака
#   короткая новость pan.am                  — 6.7% текста, 3524 знака
#   новость Caucasus Watch                   — 11.4% текста, 5212 знаков
#   большой гид по налогам                   — 14.2% текста, 16645 знаков
# По одной доле новость от пустого каркаса не отличить: 3% и 6.7% слишком
# близко. Разделяет абсолютная длина: полторы тысячи знаков — это меню и
# подвал, и ничего больше. Поэтому требуем оба признака сразу.
TEXT_SHARE = 0.05
MIN_TEXT = 2_500
MIN_HTML_BYTES = 15_000


def _looks_script_built(raw: bytes, body: str, content_type: str) -> bool:
    """Страница отдалась, но данных в ней нет — их дорисует браузер."""
    if "html" not in content_type:
        return False
    if len(raw) < MIN_HTML_BYTES:
        return False
    return len(body) < MIN_TEXT and len(body) < len(raw) * TEXT_SHARE


SHIELD_MARKERS = (
    b"just a moment",
    b"challenge-error",
    b"cf-chl",
    b"cdn-cgi/challenge-platform",
    b"performing security verification",
    b"__cf_chl",
)


def _shield_in(raw: bytes) -> bool:
    """Щит от автоматики узнаётся по телу ответа даже на коде 403."""
    lowered = raw[:6000].lower()
    return any(marker in lowered for marker in SHIELD_MARKERS)


def _looks_blocked(body: str) -> Optional[str]:
    lowered = body[:3000].lower()
    for needle, label in BOT_WALL:
        if needle in lowered:
            return label
    return None
